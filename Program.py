# make sure to pip3 install openpyxl
import openpyxl
# File GUI:
import tkinter as tk
import tkinter.filedialog as filedialog
from tkinter import ttk
from tkinter.filedialog import asksaveasfile 
from tkinter import messagebox
from tkinter import LabelFrame
from tkinter import Checkbutton
from tkinter import IntVar
from tkinter import StringVar
# CSV to Excel formatting:
import csv
import os # file extension removal
# Menu bar:
from tkinter import Menu
from tkinter import OptionMenu
#TreeView data to Excel:
import openpyxl.styles
import pandas as pd
# Number checking:
import re
#for the month conversion:
import datetime
#for nan values in zero conversion:
import numpy as np

################################
#Program Version:
VERSION_MAJOR = 1
VERSION_MINOR = 0
VERSION_PATCH = 0
################################

excel_save_file_extension = '.xlsx'
hasLoadedFileData = False # this if used to tell if we have data in the TreeView or not for saving it when opening a new file.
col_storage_pd = pd.DataFrame()
wb = None
removed_columns_str = "(REMOVED COLUMN)"
custom_name_pattern = r"^[A-Za-z].*"
allow_multiple_file_uploads_bool = False
cate_add_warnings_bool = True
custom_category_options = []
init_column_width_setting = 125
column_width_setting = 125 # this is modified later to be dynamic per the culumn header count.

r = tk.Tk()


#r.maxsize(width=1800, height=1800)

# Using treeview widget    
treev = ttk.Treeview(r, height=20, selectmode ='extended')


'''
    Setting the Select Mode:
        browse: (Default) Only one item can be selected at a time. Clicking on an item selects it and deselects any previously selected item.
        extended: Multiple items can be selected by holding down the Shift key or Ctrl key (Command key on macOS).
        none: Disables selection entirely.
'''

def get_column_index(tree, column_heading):
        for i, col in enumerate(tree["columns"]):
            if tree.heading(col)["text"] == column_heading:
                return i
        return None  # Column heading not found

def open_export_window():
    CATEGORY = 'Category'
    DATE = 'Date'
    DESCIPTION = 'Description'
    VALUE = 'Value'
    MONTH_TAG = 'month_num_head'
    YEAR_TAG = 'year_num_head'
    check_but1_var = IntVar()
    check_but2_var = IntVar()

    global remove_zero_months_var
    export_date_column = date_var.get()
    export_desc_column = desc_var.get()
    export_value_column = value_var.get()

    # Get the month to month data:
    months_df = pd.DataFrame(columns=['Month','Value'])
    category_per_month_value_df = pd.DataFrame(columns=['Category'])


    def clear_entry():
        ent_new.delete(0, tk.END)

    def reverse(x):
        return x[::-1]
    
    def calculate_custom_category_totals():
        return 0
    
    def initiate_export():
        # Check if at least one of the checkboxes are checked. Needed for export.
        if not check_but1_var.get() and not check_but2_var.get():
            messagebox.showwarning("Option Selection!", "Please verify that you have selected at least one of the options for export sheets.", parent=export_screen)
            return

        wbExport = openpyxl.Workbook()

        
        #Create list of 'id's
        listOfEntriesInTreeView=categories.get_children()

        if check_but1_var.get(): #get the state
            #export the summary
            #List of months for worksheet titles
            months = ['January', 'February', 'March', 'April', 'May', 'June', 
                    'July', 'August', 'September', 'October', 'November', 'December']

            # Create a worksheet for each month
            first_num_year = True
            first_num = True
            temp_ws = wbExport.create_sheet(title="Summary")
            for i_head, each_category in enumerate(listOfEntriesInTreeView):  
                # taking the two children: year, month we need to get the children of those and pull the numbers.
                year_or_month = categories.get_children(each_category)
                for i, year_or_month_at_index in enumerate(year_or_month):

                    if YEAR_TAG in categories.item(year_or_month_at_index)["tags"]:
                        cell_str = ""
                        cell_row_based_on_start_index = i_head + 1 + 1 # +1 because values have to start at 1 and not 0 and also want to add one more for the items not at the top.
                        # Add column headers
                        
                        cell_str = "D1"
                        temp_ws[cell_str].font = openpyxl.styles.Font(size=20, bold=True, color="000000")

                        # Align text
                        temp_ws[cell_str].alignment = openpyxl.styles.Alignment(horizontal="center")
                        temp_ws[cell_str] = str("Current Yearly Spendings")
                                                
                        cell_str = "E1"
                        temp_ws[cell_str].font = openpyxl.styles.Font(size=20, bold=True, color="000000")

                        # Align text
                        temp_ws[cell_str].alignment = openpyxl.styles.Alignment(horizontal="center")
                        temp_ws[cell_str] = str("Allocated Yearly Spendings")

                        #Now we have to add the numbers to the sheet:
                        
                        num_children = categories.get_children(year_or_month_at_index)
                        for i2 , data in enumerate(num_children):
                            if first_num_year:
                                cell_str = "D" + str(cell_row_based_on_start_index)
                                first_num_year = False
                            else:
                                cell_str = "E" + str(cell_row_based_on_start_index)
                                first_num_year = True

                            #print(categories.item(data)['values'][0])
                            

                            # Now you can apply your styles and alignment
                            # Change font style
                            temp_ws[cell_str].font = openpyxl.styles.Font(size=12, bold=False, color="03fc6f")

                            # Align text
                            temp_ws[cell_str].alignment = openpyxl.styles.Alignment(horizontal='right')
                            temp_ws[cell_str] = str(categories.item(data)['values'][0])
                    if MONTH_TAG in categories.item(year_or_month_at_index)["tags"]:
                        cell_str = ""
                        cell_row_based_on_start_index = i_head + 1 + 1 # +1 because values have to start at 1 and not 0 and also want to add one more for the items not at the top.
                        # Add column headers
                        cell_str = "A1"
                        temp_ws[cell_str].font = openpyxl.styles.Font(size=20, bold=True, color="000000")

                        # Align text
                        temp_ws[cell_str].alignment = openpyxl.styles.Alignment(horizontal="center")
                        temp_ws[cell_str] = str("Category")
                        
                        cell_str = "B1"
                        temp_ws[cell_str].font = openpyxl.styles.Font(size=20, bold=True, color="000000")

                        # Align text
                        temp_ws[cell_str].alignment = openpyxl.styles.Alignment(horizontal="center")
                        temp_ws[cell_str] = str("Current Average Monthly Spendings")
                                                
                        cell_str = "C1"
                        temp_ws[cell_str].font = openpyxl.styles.Font(size=20, bold=True, color="000000")

                        # Align text
                        temp_ws[cell_str].alignment = openpyxl.styles.Alignment(horizontal="center")
                        temp_ws[cell_str] = str("Allocated Monthly Spendings")

                        #if CATEGORY in categories.item(data)["tags"]:
                        #print(categories.item(data)['values'][0])
                        cell_str = "A" + str(cell_row_based_on_start_index)

                        # Now you can apply your styles and alignment
                        # Change font style
                        temp_ws[cell_str].font = openpyxl.styles.Font(size=16, bold=True, color="FF0000")

                        # Align text
                        temp_ws[cell_str].alignment = openpyxl.styles.Alignment(horizontal="center")
                        temp_ws[cell_str] = str(categories.item(each_category)['values'][0])

                        #Now we have to add the numbers to the sheet:
                        
                        num_children = categories.get_children(year_or_month_at_index)
                        for i2 , data in enumerate(num_children):
                            if first_num:
                                cell_str = "B" + str(cell_row_based_on_start_index)
                                first_num = False
                            else:
                                cell_str = "C" + str(cell_row_based_on_start_index)
                                first_num = True

                            #print(categories.item(data)['values'][0])
                            

                            # Now you can apply your styles and alignment
                            # Change font style
                            temp_ws[cell_str].font = openpyxl.styles.Font(size=12, bold=False, color="03fc6f")

                            # Align text
                            temp_ws[cell_str].alignment = openpyxl.styles.Alignment(horizontal='right')
                            temp_ws[cell_str] = str(categories.item(data)['values'][0])
            pass
        else:
            #do not export the summary
            pass

        if check_but2_var.get(): #get the state
            #Export month to month data
            #List of months for worksheet titles
            months = ['January', 'February', 'March', 'April', 'May', 'June', 
                    'July', 'August', 'September', 'October', 'November', 'December']

            # Create a worksheet for each month
            for month in months:
                first_num = True
                temp_ws = wbExport.create_sheet(title=month)
                for i_head, each_category in enumerate(listOfEntriesInTreeView):  
                    # taking the two children: year, month we need to get the children of those and pull the numbers.
                    year_or_month = categories.get_children(each_category)
                    for i, year_or_month_at_index in enumerate(year_or_month):

                        if MONTH_TAG in categories.item(year_or_month_at_index)["tags"]:
                            cell_str = ""
                            cell_row_based_on_start_index = i_head + 1 + 1 # +1 because values have to start at 1 and not 0 and also want to add one more for the items not at the top.
                            # Add column headers
                            #if CATEGORY in categories.item(data)["tags"]:
                            #print(categories.item(data)['values'][0])
                            cell_str = "A" + str(cell_row_based_on_start_index)

                            #merge all top cells and input text for Month
                            temp_ws["A1"].font = openpyxl.styles.Font(size=24, bold=True, color="03fc6f")
                            temp_ws["A1"].alignment = openpyxl.styles.Alignment(horizontal="center")
                            temp_ws["A1"] = str("Month of " + month + " Expenses")
                            temp_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

                            # Now you can apply your styles and alignment
                            # Change font style
                            temp_ws[cell_str].font = openpyxl.styles.Font(size=16, bold=True, color="FF0000")

                            # Align text
                            temp_ws[cell_str].alignment = openpyxl.styles.Alignment(horizontal="center")
                            temp_ws[cell_str] = str(categories.item(each_category)['values'][0])

                            #Now we have to add the numbers to the sheet:
                            
                            num_children = categories.get_children(year_or_month_at_index)
                            for i2 , data in enumerate(num_children):
                                if first_num:
                                    cell_str = "B" + str(cell_row_based_on_start_index)
                                    first_num = False
                                else:
                                    cell_str = "C" + str(cell_row_based_on_start_index)
                                    first_num = True

                                #print(categories.item(data)['values'][0])
                                

                                # Now you can apply your styles and alignment
                                # Change font style
                                temp_ws[cell_str].font = openpyxl.styles.Font(size=12, bold=False, color="03fc6f")

                                # Align text
                                temp_ws[cell_str].alignment = openpyxl.styles.Alignment(horizontal='right')
                                temp_ws[cell_str] = str("0") #default set month value to 0
                                for month_data in categories.get_children(data):
                                    find_val = str(categories.item(month_data)['text']).find(month)
                                    #print(find_val)
                                    #print(str(categories.item(month_data)['text']))
                                    if find_val > -1:
                                        temp_ws[cell_str] = str(categories.item(month_data)['values'][0])
                        if YEAR_TAG:
                            pass                    
        else:
            #do not export the monthly
            pass


        temp_ws_transactions = wbExport.create_sheet(title="Transactions")

        for index_of_item, full_list_row in enumerate(treev.get_children()):
            #We need to add the code for going through the whole data set and adding up all the values labeled in a given custom column.
            # but before we do that we need to make the value column essential along with maybe the date and description column. So make them all need to not be null in value.
            #try:
            index_date = get_column_index(treev, date_var.get())
            index_desc = get_column_index(treev, desc_var.get())
            index_value = get_column_index(treev, value_var.get())

            # get data:
            category_data = treev.item(full_list_row)['values'][0]
            date_data = treev.item(full_list_row)['values'][index_date]
            desc_data = treev.item(full_list_row)['values'][index_desc]
            value_data = treev.item(full_list_row)['values'][index_value]
                    
            temp_ws_transactions[str("A"+str(index_of_item+1))] = desc_data
            temp_ws_transactions[str("B"+str(index_of_item+1))] = category_data
            temp_ws_transactions[str("C"+str(index_of_item+1))] = date_data
            temp_ws_transactions[str("D"+str(index_of_item+1))] = value_data
        

        # Remove the default sheet
        del wbExport['Sheet']

        # Save the workbook
        wbExport.save("Expenses.xlsx")
        print("Workbook created with a sheet for each month.")


    def edit():

        #Check if we are modifying the numbers and if the input is a number:
        # Get selected item to Edit
        selected_item = categories.selection()[0]
        text = categories.item(selected_item)['text']
        value = categories.item(selected_item)['values'][0]

        #reverse the text and values so that we can make sure to replace the end of the data and not the beginning.
        r_text = reverse(str(text))
        r_value = reverse(str(value))
        new_text = r_text.replace(str(r_value), reverse(ent_new.get()), 1)

        if VALUE in categories.item(selected_item)["tags"]:
            if not is_number_with_two_decimals(ent_new.get()):
                messagebox.showerror("Bad Input!", "Please verify you are entering a number that is a valid monetary value. That is a number with no more than 2 decimal places.")
                return
            else:
                categories.item(selected_item, text=reverse(new_text), values=(ent_new.get()))
                clear_entry()
        if CATEGORY in categories.item(selected_item)["tags"]:
            if len(ent_new.get()) == 0 or not re.match(custom_name_pattern, ent_new.get()):
                messagebox.showerror("Bad Input!", "Please verify you are entering a valid non-blank custom category name which at least starts with one letter.")
                return
            else:

                categories.item(selected_item, text=reverse(new_text), values=(ent_new.get()))

                #update the custom name in the whole tree:
                for row in treev.get_children():
                    if value == treev.item(row)['values'][0]:
                        old_values = treev.item(row)['values']
                        old_values.pop(0) # this will remove the old custom name
                        new_values = (ent_new.get(),) + tuple(old_values)
                        treev.item(row, values=new_values)
                        #update the custom name in drop down menu
                # Get the current values
                values = list(dropdown["values"])

                # Replace a specific value
                i = values.index(str(value))
                values[i] = ent_new.get()

                # Update the Combobox with the modified values
                dropdown["values"] = values
                dropdown.set("")
                clear_entry()


    export_screen = tk.Toplevel(r)
    export_screen.title("Export")
    export_screen.minsize(width=400,height=300)
    # Set window position (x, y)
    export_screen.geometry("+150+150") # window position

    #Force this window to be at top and stay on top:
    export_screen.attributes("-topmost", True)

    # Add widgets to the new window here
    tk.Label(export_screen, text="These are the export options").pack()
    tk.Button(export_screen, text="Export", command=initiate_export).pack(side="bottom", anchor="se", padx=10, pady=10)

    categories = ttk.Treeview(export_screen, height=20, selectmode ='browse') #, columns=("Budget")
    categories.heading("#0", text='Budget Data')
    categories.column('#0', minwidth=100, width=300, stretch=True)
    categories.pack(side="left")

    # Get the number of columns
    test = categories["columns"]
    num_columns = len(categories["columns"])

    unique_values = []
    monthly_value = []
    yearly_value = 0

    df = pd.DataFrame(columns = ['Category' , 'Date', 'Description' , 'Value'])

    #safty net:
    if len(treev.get_children()) == 0:
        
        messagebox.showerror("No Values Imported", "Please Import some data.", parent=export_screen)
        #so not to continue we return and force this to fail.
        export_screen.destroy()
        return

    head_row = treev.get_children()[0]

    for row in treev.get_children():
        #We need to add the code for going through the whole data set and adding up all the values labeled in a given custom column.
        # but before we do that we need to make the value column essential along with maybe the date and description column. So make them all need to not be null in value.
        try:
            index_date = get_column_index(treev, date_var.get())
            index_desc = get_column_index(treev, desc_var.get())
            index_value = get_column_index(treev, value_var.get())

            # get data:
            category_data = treev.item(row)['values'][0]
            date_data = treev.item(row)['values'][index_date]
            desc_data = treev.item(row)['values'][index_desc]
            value_data = treev.item(row)['values'][index_value]
        
        except TypeError:
            messagebox.showerror("Select Column Data", "Please verify you have the corresponding columns selected for Date, Description and, Value.", parent=export_screen)
            #so not to continue we return and force this to fail.
            export_screen.destroy()
            return

        #now place the values into the tree based on their index positions in the treeview.
        df.loc[len(df)] = [category_data, date_data, desc_data, value_data]
        pass


    months = ['January', 'February', 'March', 'April', 'May', 'June', 
                    'July', 'August', 'September', 'October', 'November', 'December']
    months_value = [0,0,0,0,0,
                    0,0,0,0,0,
                    0,0]
    cat_list = []

    for distinguishable in pd.unique(df['Category']):
        new_row = pd.DataFrame({'Category':[distinguishable]})
        for in_i, mon in enumerate(months):
            new_row[mon] = months_value[in_i]
        category_per_month_value_df = pd.concat([category_per_month_value_df, new_row], sort=False, ignore_index=True)

    #print(category_per_month_value_df)

    '''
    for distinguishable in pd.unique(df['Category']):
        new_row = pd.DataFrame({'Category':[distinguishable],'Month_Values_List':[months_value]})
        #category_per_month_value_df[len(category_per_month_value_df)] = new_row
        category_per_month_value_df = pd.concat([category_per_month_value_df, new_row], sort=False, ignore_index=True)

    print(category_per_month_value_df)
    '''
    
    # Convert the 'Date' column to datetime format
    df['Date'] = pd.to_datetime(df['Date'], format='mixed')
    #add the month column to the dataframe:
    df['Month'] = df['Date'].dt.month_name()
    
    for index, row in df.iterrows():
        for month in months:
            d_a_t_e = row['Month']
            if month == row['Month']:
                #category_per_month_value_df[row['Category']][]
                row_index = category_per_month_value_df.index.get_loc(category_per_month_value_df[category_per_month_value_df['Category'] == row['Category']].index[0])
                category_per_month_value_df.loc[row_index, row['Month']] = category_per_month_value_df.loc[row_index, row['Month']] + float(row['Value'])
                # category_per_month_value_df[row_index][months.index(row['Month'])] = category_per_month_value_df[row_index][months.index(row['Month'])] + float(row['Value'])
                break
    '''
    for i,month in enumerate(months):
        #new_row = pd.DataFrame({month:months_value[i]})
        # Converting to the dataframe 
        new_row = pd.DataFrame({'Month': [month],'Value': [months_value[i]]})
        months_df.loc[len(months_df)] = [month,months_value[i]]
        #pd.concat([months_df,new_row], ignore_index=True)
    '''

    def average_row_values(df, exclude_cols):
        """
        Averages values in each row of a DataFrame, excluding specified columns and zero values.

        Args:
            df (pd.DataFrame): The input DataFrame.
            column category to output
            exclude_cols (list): A list of column names to exclude from the average.

        Returns:
            the average of the column category row.
        """
        df_filtered = df.drop(columns=exclude_cols)
        
        # Replace 0 with NaN to exclude them from the mean calculation
        df_filtered = df_filtered.replace(0, np.nan)

        row_average = df_filtered.mean(axis=1, skipna=True)
        return row_average
    #print("DF:",df)
    #print("CATE:",category_per_month_value_df)
    for index, unique in enumerate(pd.unique(df['Category'])):
        if unique not in unique_values and not unique == "":
            # Calculate the average of rows ignoring 0 values
            # Columns to exclude from the sum
            # Replace 0 with NaN to exclude them from the sum
            # Row index to sum (e.g., row with index 1)
            row_index = category_per_month_value_df.index.get_loc(category_per_month_value_df[category_per_month_value_df['Category'] == unique].index[0])

            # Columns to exclude (e.g., 'col1' and 'col3')
            exclude_cols = ['Category']

            # drop category from the list of month values:
            row = category_per_month_value_df.loc[row_index].drop(exclude_cols)
            
            #For calculations that need the zero values we will create a seperate dataframe:
            row_with_zero = row

            def find_last_nonzero(df_temp, exclude_cols):
                '''
                    This will output nan if no values other than zero found
                '''
                df_filtered = df_temp.drop(columns=exclude_cols)
                results = []
                for index_temp in reversed(df_filtered.index):
                    row_temp = df_filtered.loc[index_temp]
                    first_nonzero_index = -1
                    for i_temp in range(0,len(row_temp)):
                        temp_value_out = row_temp.iloc[i_temp]
                        if row_temp.iloc[i_temp] != 0:
                            results.append(i_temp)
                            #print(results)
                            #break
                    #if first_nonzero_index != -1:
                    #    results.append(first_nonzero_index)
                    if first_nonzero_index == -1:
                        results.append(np.nan)
                return int(np.nanmax(results)) # Restore original order [::-1]
            
            #intially set row_values so it has a value:
            row_values = category_per_month_value_df.loc[row_index].drop(exclude_cols)

            lastMonth = 11 #because we have 0 included in 0, 1, 2,... 11
            if include_zeros_month_avg_var.get():
                dropAfterIndex = find_last_nonzero(category_per_month_value_df,exclude_cols)
                #print(dropAfterIndex)
                if not dropAfterIndex == np.nan:
                    lastMonth = dropAfterIndex
                    row_values = category_per_month_value_df.loc[row_index].drop(exclude_cols).iloc[:dropAfterIndex+1] #drop columns after last data value index.
                else:
                    # we drop all columns except one because the average is gonna be zero.
                    row_values = category_per_month_value_df.loc[row_index].drop(exclude_cols).iloc[:, :0+1] #drop columns after last data value index.
                #row_values = row_values.loc[row_index].drop(exclude_cols)
            else:
                # Replace 0 with NaN to exclude them from the sum
                row_values = row_values.replace(0, pd.NA).dropna()

                # Get the row, exclude columns, and replace 0s with NaN
                row_values = category_per_month_value_df.loc[row_index].drop(exclude_cols).replace(0, np.nan).infer_objects(copy=False)

            # Calculate the mean, ignoring NaNs
            #print("ROW", row)
            #print("ROW_VALUES", row_values)
            average = round(row_values.mean(), 2)

            # Calculate the sum
            row_sum = row_values.sum()
            year_current = round(row_sum, 2)

            monthly_string = "Budget Current Average: $" + str(average)
            month_string_1st_half = "Budget "
            month_string_2nd_half = " Current: $"
            yearly_string = "Budget Current: $" + str(year_current)
            unique_values.append(unique)
            str_1_head = "Category Name: " + unique
            head_1 = categories.insert("", 'end', text =str_1_head, values=(unique), tags=CATEGORY)
            child_head_m = categories.insert(head_1, 'end', text ="Monthly", values=("0"), tags=MONTH_TAG)
            child_head_y = categories.insert(head_1, 'end', text ="Yearly", values=("0"), tags=YEAR_TAG)
            child_m1 = categories.insert(child_head_m, 'end', text =monthly_string, values=(str(average)), tags=VALUE)
            #for the value in each month:
            # Get the index (row header text) for all months
            row_headers = row_with_zero.index
            # Convert the index to a list of all months
            row_headers_list = list(row_headers)
            for month_index, month_value in enumerate(row_with_zero): # This is category_per_month_value_df without the category column and with the specific category totals per month.
                if remove_zero_months_var.get(): # do not add months that are 0 in value to the month average list.
                    if month_value == 0:
                        continue
                if include_zeros_month_avg_var.get(): # do not add months that are after the final month to the average.
                    if lastMonth < month_index:
                        continue
                temp_string = month_string_1st_half + row_headers_list[month_index] + month_string_2nd_half + str(month_value)
                categories.insert(child_m1, 'end', text =temp_string, values=(str(month_value)), tags=VALUE)
            child_m2 = categories.insert(child_head_m, 'end', text ="Budget Allocated: $0", values=("0"), tags=VALUE)
            child_y1 = categories.insert(child_head_y, 'end', text =yearly_string, values=(str(year_current)), tags=VALUE)
            child_y2 = categories.insert(child_head_y, 'end', text ="Budget Allocated: $0", values=("0"), tags=VALUE)


    # This will create a LabelFrame
    label_export_options_1 = LabelFrame(export_screen, text='Export Pages', font = "50")
    label_export_options_1.pack(side='top', anchor='n')
    # This will create a LabelFrame
    #label_export_sub_1 = LabelFrame(label_export_options_1, text='Create Summary Page', font = "20")
    #label_export_sub_1.pack(side="left")
    # This will create a LabelFrame
    #label_export_sub_2 = LabelFrame(label_export_options_1, text='Create Monthly Pages', font = "20")
    #label_export_sub_2.pack(side="right")
    check_but1 = tk.Checkbutton(label_export_options_1, text="Create Summary Page", 
                        variable = check_but1_var, 
                        onvalue = 1, 
                        offvalue = 0 
                        #height = 2, 
                        #width = 20,
                        )
    check_but2 = tk.Checkbutton(label_export_options_1, text="Create Monthly Pages", 
                        variable = check_but2_var, 
                        onvalue = 1, 
                        offvalue = 0 
                        #height = 2, 
                        #width = 20,
                        )
    check_but1.pack()
    check_but2.pack()

    '''# This will create a LabelFrame
    label_export_options_2 = LabelFrame(export_screen, text='Export Options 2', font = "50")
    label_export_options_2.pack(side='top', anchor='n')
    # This will create a LabelFrame
    label_export_sub_3 = LabelFrame(label_export_options_2, text='Export sub 1', font = "20")
    label_export_sub_3.pack(side="left")
    # This will create a LabelFrame
    label_export_sub_4 = LabelFrame(label_export_options_2, text='Export sub 2', font = "20")
    label_export_sub_4.pack(side="right")
    check_but3 = tk.Checkbutton(label_export_sub_3, text="test1")
    check_but4 = tk.Checkbutton(label_export_sub_4, text="test2")
    check_but3.pack()
    check_but4.pack()'''

    # This will create a LabelFrame
    label_export_options_3 = LabelFrame(export_screen, text='Change Custom Category Name or Min/Max Value', font = "50")
    label_export_options_3.pack(side='top', anchor='n')
    # This will create a LabelFrame
    label_export_sub_5 = LabelFrame(label_export_options_3, text='Export sub 1', font = "20")
    label_export_sub_5.pack(side="left")
    # This will create a LabelFrame
    label_export_sub_6 = LabelFrame(label_export_options_3, text='Export sub 2', font = "20")
    label_export_sub_6.pack(side="right")
    ent_new = tk.Entry(label_export_options_3)
    but1 = tk.Button(label_export_options_3, text="Change", command=edit)
    ent_new.pack()
    but1.pack()

    '''
    check_but5 = tk.Checkbutton(label_export_sub_5, text="test1")
    check_but6 = tk.Checkbutton(label_export_sub_6, text="test2")
    check_but5.pack()
    check_but6.pack()
    '''

    # Get the number of columns
    test = categories["columns"]
    num_columns = len(categories["columns"])

def find_and_change_date_item(old_item, new_item):
    current_values = list(dropdown_date['values'])
    try:
        index = current_values.index(old_item)
        current_values[index] = new_item
        dropdown_date['values'] = tuple(current_values) 
    except ValueError:
        print(f"{old_item} not found in Combobox")

def find_and_change_desc_item(old_item, new_item):
    current_values = list(dropdown_desc['values'])
    try:
        index = current_values.index(old_item)
        current_values[index] = new_item
        dropdown_desc['values'] = tuple(current_values) 
    except ValueError:
        print(f"{old_item} not found in Combobox")

def find_and_change_value_item(old_item, new_item):
    current_values = list(dropdown_value['values'])
    try:
        index = current_values.index(old_item)
        current_values[index] = new_item
        dropdown_value['values'] = tuple(current_values) 
    except ValueError:
        print(f"{old_item} not found in Combobox")

def find_and_change_sign_item(old_item, new_item):
    current_values = list(change_sign_combobox['values'])
    try:
        index = current_values.index(old_item)
        current_values[index] = new_item
        change_sign_combobox['values'] = tuple(current_values) 
    except ValueError:
        print(f"{old_item} not found in Combobox")


def change_sign_item():
    try:
        #get column selected:
        selected_items = treev.selection()
        for selected_item in selected_items:
            #get row:
            item_data = treev.item(selected_item)['values']
            
            # get column index:
            box = change_sign_combobox.get()

            index_of_col_name = get_column_index(treev,box)
            if index_of_col_name == None:
                print("Failed to find value")

            # get value:
            cell_value = item_data[index_of_col_name]

            #flip value:
            #iid = selected_row[index_of_col_name]
            

            # Modify the values (example: change the first column)
            item_data[index_of_col_name] = str(float(cell_value) * -1)

            treev.item(selected_item, values=tuple(item_data))
            #update_treeview_cell(iid=iid, column=change_column_combobox.get(), new_value=)

    except ValueError:
        print(f"item not found in Combobox")

def find_and_change_checkbox_item(old_item, new_item):
    for i, button in enumerate(ButtonsN):
        if button['text'] == old_item:
            try:
                for i_in, header in enumerate(treev['columns']):
                    if header == old_item:
                        col_txt = '#' + str(i_in)
                        #treev.heading(col_txt, text=new_item)
                print(i)
                ButtonsN[i].config(text=new_item)
                #treev.heading('old_item', text='Column One')
                
                #print(ButtonsN[i])
            except ValueError:
                print(f"{old_item} not found in Combobox")

def find_and_change_checkbox_item(old_item, new_item):
    for i, button in enumerate(ButtonsN):
        if button['text'] == old_item:
            try:
                for i_in, header in enumerate(treev['columns']):
                    if header == old_item:
                        col_txt = '#' + str(i_in)
                        #treev.heading(col_txt, text=new_item)
                print(i)
                ButtonsN[i].config(text=new_item)
                #treev.heading('old_item', text='Column One')
                
                #print(ButtonsN[i])
            except ValueError:
                print(f"{old_item} not found in Combobox")
                

def column_name_change():
    if len(change_column_entry.get()) > 0:
        old_column = change_column_combobox.get()
        new_column = change_column_entry.get()
        values_col = list(change_column_combobox['values'])
        index = values_col.index(old_column)
        values_col.pop(index)
        values_col.append(new_column)
        change_column_combobox['values'] = values_col
        change_column_combobox.set(new_column)
        old_column_index = get_column_index(treev, old_column)
        if old_column_index >= 0:
            #treev.heading(old_column_index, text=new_column)
            find_and_change_date_item(old_column, new_column)
            find_and_change_desc_item(old_column, new_column)
            find_and_change_value_item(old_column, new_column)
            find_and_change_checkbox_item(old_column, new_column)
            find_and_change_sign_item(old_column, new_column)
            
            cols = list(treev['columns'])
            print(cols)
            cols[index+1] = new_column
            treev.configure(columns=tuple(cols))
            treev
            for i_col, column_val in enumerate(cols):
                # Assigning the heading names to the respective columns
                treev.column(i_col, width=column_width_setting, stretch=False)
                treev.heading(column=i_col ,text=column_val)
            #treev.configure(width=(column_width_setting*len(list(treev['columns']))))
            treev['show'] = 'headings'
            print(cols)
        
        


# Adjust size
# I want it resizable so that the scrolling and the data boxes are not messed up by it.
r.minsize(width=600,height=10)
# Set window position (x, y)
r.geometry("+100+100") # window position
r.resizable(width=False, height=False)
r.title('Budgetting App')

#this covers all Labels:
data_options_label = LabelFrame(r, text='Data Options', font = "30")
data_options_label.pack(side='left', fill='both')

#category grouping label
label_left_data_options = LabelFrame(data_options_label, borderwidth=0)
label_change_tree_group = LabelFrame(label_left_data_options, borderwidth=0)
lebel_change_column = LabelFrame(label_change_tree_group, text='Update Column Name', font=35)
lebel_change_sign = LabelFrame(label_change_tree_group, text='Update Value Sign +/-',font=35)
change_column_entry_label = ttk.LabelFrame(lebel_change_column, text='Choose Column to Change')
change_column_entry = ttk.Entry(lebel_change_column)
change_column_button = ttk.Button(lebel_change_column, text='Change Column Name', command=column_name_change)
change_sign_combobox_label = ttk.LabelFrame(lebel_change_sign, text='Choose Column to Change')
change_sign_button = ttk.Button(lebel_change_sign, text='Change +/- Sign', command=change_sign_item)

label_left_data_options.pack(side='left')
label_change_tree_group.pack(side='top', anchor='n', fill='both', pady=+30)
lebel_change_column.pack(side='left', anchor='center')
lebel_change_sign.pack(side='right', anchor='center')
change_column_entry_label.pack()
change_sign_combobox_label.pack()
change_column_entry.pack()
change_column_button.pack()
change_sign_button.pack()

#category grouping label
label_category_group = LabelFrame(label_left_data_options, borderwidth=0)
label_category_group.pack(side='bottom',anchor='s', pady=+30)

# This will create a LabelFrame to hold (Date, Description, Value)
label_frame_date_desc_value = LabelFrame(label_category_group, text='Corresponding Columns', font = "30", pady=30)
label_frame_date = LabelFrame(label_frame_date_desc_value, text='Date Column', font = "20")
label_frame_desc = LabelFrame(label_frame_date_desc_value, text='Description Column', font = "20")
label_frame_value = LabelFrame(label_frame_date_desc_value, text='Value Column', font = "20")



# This will create a LabelFrame
label_frame = LabelFrame(data_options_label, text='Remove Columns', font = "50")

# This will create a LabelFrame
label_category_selection = LabelFrame(label_category_group, text='Custom Categories', font = "50")
# This will create a LabelFrame
label_category_selection_add = LabelFrame(label_category_selection, text='Add Custom Category', font = "20")
# This will create a LabelFrame
label_category_selection_select = LabelFrame(label_category_selection, text='Set Row Category', font = "20")

ButtonsN = []

# List to hold the checkboxes' variable references
check_vars = []



def get_selected_tree_row_add_category():
    #if there have been no values added to category.
    if len(dropdown["values"]) == 0:
        return
    selected_items = treev.selection()

    for item in selected_items:
        values_in_row = treev.item(item).get("values")
        '''new_values = []
        item_1 = True
        for values in values_in_row:
            if item_1:
                new_values.append(dropdown.current())
            else:
                new_values.append(values)'''
        
        treev.set(item=item,column=0,value=dropdown["values"][dropdown.current()])



# Constructing vertical scrollbar
# with treeview
verscrlbar = ttk.Scrollbar(r, 
                           orient ="vertical", 
                           command = treev.yview)
# Constructing horizontal scrollbar
# with treeview
horzscrlbar = ttk.Scrollbar(r, 
                           orient ="horizontal", 
                           command = treev.xview)

# Calling pack method w.r.to vertical 
# scrollbar
verscrlbar.pack(side ='right', fill ='y', anchor='ne')

horzscrlbar.pack(side='bottom', fill='x', anchor='sw')

# Calling pack method w.r.to treeview
treev.pack(side ='left', fill='x', anchor='nw')

label_frame_date_desc_value.pack(side ='top')

label_frame_date.pack()
label_frame_desc.pack()
label_frame_value.pack()

label_frame.pack(side='left', anchor='e')

label_category_selection.pack(side='top')

label_category_selection_add.pack(side='left')
label_category_selection_select.pack(side='right')

#adding stuff to the custom categories
entry = ttk.Entry(label_category_selection_add, width = 20)
entry.pack()



def clear_new_category_entry():
        entry.delete(0, tk.END)

def add_new_category():
    if not re.match(custom_name_pattern, entry.get()):
        messagebox.showerror("Bad Input!", "Please verify you are entering a valid non-blank custom category name which at least starts with one letter.")
        return
    message_str = "Would you like to add the new Category \"" + entry.get() + "\" to your column category options?"
    if messagebox.askokcancel("Add New Category", message_str) if cate_add_warnings_bool else True:
        values = list(dropdown["values"])
        dropdown["values"] = values + [entry.get()]
        clear_new_category_entry()

sign_name_var = tk.StringVar()
sign_name_prev = ""
column_name_var = tk.StringVar()
column_name_prev = ""
date_var = tk.StringVar()
date_prev = ""
desc_var = tk.StringVar()
desc_prev = ""
value_var = tk.StringVar()
value_prev = ""

date_add_toggle = []
desc_add_toggle = []
value_add_toggle = []
column_name_toggle = []
sign_name_toggle = []

def search_by_first_element(list_of_lists, target): #returns the boolean value associated with the target. if item is not found returns false.
    #print(list_of_lists)
    list_of_lists = list(list_of_lists)
    for lis in list_of_lists:
        if lis[0] == target:
            return lis[1]
    return None

def add_item_toggle_list(list_of_lists, target):
    #print(list_of_lists)
    list_of_lists = list(list_of_lists)
    list_of_lists.append([target,False]) # start false since it is gonna get flipped
    return list_of_lists

def flip_bool_tuple(list_of_lists, target):
    for i in range(len(list_of_lists)):
        if list_of_lists[i][0] == target:
            list_of_lists[i][1] = not list_of_lists[i][1]
    return list_of_lists

def modify_row(index):
    selected_items = treev.selection()
    if not selected_items:
        return  # No row selected

    selected_iid = selected_items[index]
    current_values = list(treev.item(selected_iid, 'values'))

    # Modify the values (example: change the first column)
    current_values[index] = "Updated Value"

    treev.item(selected_iid, values=tuple(current_values))

def update_columns_in_dropdowns(value_to_remove):
    global column_name_prev
    global column_name_toggle
    #toggle the add and remove
    evaluation_of_toggle = search_by_first_element(column_name_toggle, value_to_remove)
    if evaluation_of_toggle:
        
        #if we are supposed to add. Then we add the value back into the date list

        #update the toggle boolean:
        column_name_toggle = flip_bool_tuple(column_name_toggle, value_to_remove)

        #first get the complete list:
        column_options = list(change_column_combobox['values'])

        #then add the value to the list:
        column_options.append(value_to_remove) 

        #now update the new list:
        change_column_combobox['values'] = column_options
    else:
        #if we are supposed to remove. Then we remove the value from the date list#if we are supposed to remove. Then we remove the value from the date list

        if evaluation_of_toggle is None:
            #This means that value does not exist and we need to add it to the toggle list:
            column_name_toggle = add_item_toggle_list(column_name_toggle, value_to_remove)

        #update the toggle boolean:
        column_name_toggle = flip_bool_tuple(column_name_toggle, value_to_remove)

        #first get the complete list:
        column_options = list(change_column_combobox['values'])

        #first check if a value is selected:
        #remove selected value:
        try:
            if len(date_var.get()) > 0: #this means it is not blank
                #check if the value is the column selected for deletion:
                if value_to_remove == column_name_var.get():
                    change_column_combobox.set('')
        except UnboundLocalError:
            #This means that the value date_var was empty ''
            pass

        #then add the value to the list:
        try:
            column_options.remove(value_to_remove) 
        except ValueError:
            pass

        #now update the new list:
        change_column_combobox['values'] = column_options
        pass

    return # so not to execute the rest of the function.


def update_date_dropdown(event):
    global date_prev
    global date_add_toggle
    if isinstance(event,str): # if is it s string then we know it is the column we need to remove
        #toggle the add and remove
        evaluation_of_toggle = search_by_first_element(date_add_toggle, event)
        if evaluation_of_toggle:
            #if we are supposed to add. Then we add the value back into the date list

            #update the toggle boolean:
            date_add_toggle = flip_bool_tuple(date_add_toggle, event)

            #first get the complete list:
            date_options = list(dropdown_date['values'])

            #then add the value to the list:
            date_options.append(event) 

            #now update the new list:
            dropdown_date['values'] = date_options

            #finally update the selected item to nothing:

        else:
            #if we are supposed to remove. Then we remove the value from the date list

            if evaluation_of_toggle is None:
                #This means that value does not exist and we need to add it to the toggle list:
                date_add_toggle = add_item_toggle_list(date_add_toggle, event)

            #update the toggle boolean:
            date_add_toggle = flip_bool_tuple(date_add_toggle, event)

            #first get the complete list:
            date_options = list(dropdown_date['values'])

            #first check if a value is selected:
            #remove selected value:
            try:
                if len(date_var.get()) > 0: #this means it is not blank
                    #check if the value is the column selected for deletion:
                    if event == date_var.get():
                        dropdown_date.set('')
            except UnboundLocalError:
                #This means that the value date_var was empty ''
                pass

            #then add the value to the list:
            try:
                date_options.remove(event) 
            except ValueError:
                pass

            #now update the new list:
            dropdown_date['values'] = date_options
            pass

        return # so not to execute the rest of the function.
    if(len(date_var.get()) > 0 and not date_var.get() == date_prev):
        try:
            desc_options = list(dropdown_desc['values'])
            desc_options.remove(date_var.get())
            if not date_prev == "": desc_options.append(date_prev) 
            dropdown_desc['values'] = desc_options
        except ValueError:
            pass
        try:
            value_options = list(dropdown_value['values'])
            value_options.remove(date_var.get())
            if not date_prev == "": value_options.append(date_prev) 
            dropdown_value['values'] = value_options
        except ValueError:
            pass
        
        date_prev = date_var.get()
def update_desc_dropdown(event):
    global desc_prev
    global desc_add_toggle
    if isinstance(event,str): # if is it s string then we know it is the column we need to remove
        #toggle the add and remove
        evaluation_of_toggle = search_by_first_element(desc_add_toggle, event)
        if evaluation_of_toggle:
            #if we are supposed to add. Then we add the value back into the date list

            #update the toggle boolean:
            desc_add_toggle = flip_bool_tuple(desc_add_toggle, event)

            #first get the complete list:
            desc_options = list(dropdown_desc['values'])

            #then add the value to the list:
            desc_options.append(event) 

            #now update the new list:
            dropdown_desc['values'] = desc_options

            #finally update the selected item to nothing:
            
        else:
            #if we are supposed to remove. Then we remove the value from the date list

            if evaluation_of_toggle is None:
                #This means that value does not exist and we need to add it to the toggle list:
                desc_add_toggle = add_item_toggle_list(desc_add_toggle, event)

            #update the toggle boolean:
            desc_add_toggle = flip_bool_tuple(desc_add_toggle, event)

            #first get the complete list:
            desc_options = list(dropdown_desc['values'])

            #first check if a value is selected:
            #remove selected value:
            try:
                if len(desc_var.get()) > 0: #this means it is not blank
                    #check if the value is the column selected for deletion:
                    if event == desc_var.get():
                        dropdown_desc.set('')
            except UnboundLocalError:
                #This means that the value desc_var was empty ''
                pass

            #then add the value to the list:
            try:
                desc_options.remove(event) 
            except ValueError:
                pass

            #now update the new list:
            dropdown_desc['values'] = desc_options
            pass

        return # so not to execute the rest of the function.
    if(len(desc_var.get()) > 0 and not desc_var.get() == desc_prev):
        try:
            date_options = list(dropdown_date['values'])
            date_options.remove(desc_var.get())
            if not desc_prev == "": date_options.append(desc_prev) 
            dropdown_date['values'] = date_options
        except ValueError:
            pass
        try:
            value_options = list(dropdown_value['values'])
            value_options.remove(desc_var.get())
            if not desc_prev == "": value_options.append(desc_prev) 
            dropdown_value["values"] = value_options
        except ValueError:
            pass

        desc_prev = desc_var.get()
def update_value_dropdown(event):
    global value_prev
    global value_add_toggle
    if isinstance(event,str): # if is it s string then we know it is the column we need to remove
        #toggle the add and remove
        evaluation_of_toggle = search_by_first_element(value_add_toggle, event)
        if evaluation_of_toggle:
            #if we are supposed to add. Then we add the value back into the date list

            #update the toggle boolean:
            value_add_toggle = flip_bool_tuple(value_add_toggle, event)

            #first get the complete list:
            value_options = list(dropdown_value['values'])

            #then add the value to the list:
            value_options.append(event) 

            #now update the new list:
            dropdown_value['values'] = value_options

            #finally update the selected item to nothing:
            
        else:
            #if we are supposed to remove. Then we remove the value from the date list

            if evaluation_of_toggle is None:
                #This means that value does not exist and we need to add it to the toggle list:
                value_add_toggle = add_item_toggle_list(value_add_toggle, event)

            #update the toggle boolean:
            value_add_toggle = flip_bool_tuple(value_add_toggle, event)

            #first get the complete list:
            value_options = list(dropdown_value['values'])

            #first check if a value is selected:
            try:
                if len(value_var.get()) > 0: #this means it is not blank
                    #check if the value is the column selected for deletion:
                    if event == value_var.get():
                        dropdown_value.set('')
            except UnboundLocalError:
                #This means that the value value_var was empty ''
                pass

            #then add the value to the list:
            try:
                value_options.remove(event)
            except ValueError:
                pass

            #now update the new list:
            dropdown_value['values'] = value_options
            pass

        return # so not to execute the rest of the function.
    if(len(value_var.get()) > 0 and not value_var.get() == value_prev):
        try:
            desc_options = list(dropdown_desc['values'])
            desc_options.remove(value_var.get())
            if not value_prev == "": desc_options.append(value_prev) 
            dropdown_desc["values"] = desc_options
        except ValueError:
            pass
        try:
            date_options = list(dropdown_date['values'])
            date_options.remove(value_var.get())
            if not value_prev == "": date_options.append(value_prev) 
            dropdown_date["values"] = date_options
        except ValueError:
            pass

        value_prev = value_var.get()



change_column_combobox = ttk.Combobox(change_column_entry_label, width = 20, state="readonly", textvariable=column_name_var)
change_sign_combobox = ttk.Combobox(change_sign_combobox_label, width = 20, state="readonly", textvariable=sign_name_var)
change_column_combobox.pack()
change_sign_combobox.pack()

entry_b = ttk.Button(label_category_selection_add, text='Create New Label', width = 20, command=add_new_category)
entry_b.pack()
dropdown = ttk.Combobox(label_category_selection_select, width = 20, state="readonly")
dropdown.pack()
dropdown_date = ttk.Combobox(label_frame_date, width = 12, state="readonly", textvariable=date_var)
dropdown_date.bind('<<ComboboxSelected>>', update_date_dropdown)
dropdown_date.pack()
dropdown_desc = ttk.Combobox(label_frame_desc, width = 12, state="readonly", textvariable=desc_var)
dropdown_desc.bind('<<ComboboxSelected>>', update_desc_dropdown)
dropdown_desc.pack()
dropdown_value = ttk.Combobox(label_frame_value, width = 12, state="readonly", textvariable=value_var)
dropdown_value.bind('<<ComboboxSelected>>', update_value_dropdown)
dropdown_value.pack()
dropdown_add = ttk.Button(label_category_selection_select, text='Label Selected Row', width = 20, command=get_selected_tree_row_add_category)
dropdown_add.pack()
 
# Configuring treeview
treev.configure(yscrollcommand = verscrlbar.set, xscrollcommand=horzscrlbar.set)

# The above packing was done in this order to help with the window layout. Other packing orders have not yeilded this good of a result.

def hide_object(ob):
    if type(ob) is not list:
        ob.pack_forget()
        return
    for ite in ob:
        ite.pack_forget()
def show_object(ob, appen, fill):
    if type(ob) is not list:
        ob.pack(expand = True, fill=fill, side=appen)
        return
    for ite in ob:
        ite.pack(expand = True, fill=fill, side=appen)
def open_file():
    file_path = filedialog.askopenfilename(filetypes=[('Excel', '*.xlsx'),('Excel', '*.xlsm'),('Excel', '*.xlsb'),('Excel', '*.xltx'),('Excel','*.xls'), ('CSV files', '*.csv')])
    if file_path:
        print("Selected file:", file_path)
        return file_path
    else:
        print("No file selected.")
        return False
    

def is_decimal_string(input):
    import re

    pattern = r"-?\d+\.\d+"

    match = re.search(pattern, input)

    if match:
        return True
    else:
        return False

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False
    
def is_number_with_two_decimals(string):
    pattern = r"^-?\d+(\.\d{1,2})?$"
    return bool(re.match(pattern, string))

def convert_csv_to_excel(fileName): #returns the new file name
    global excel_save_file_extension
    wb = openpyxl.Workbook()
    ws = wb.active
    with open(file=fileName, mode='r') as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            # Cell to modify
            for i in range(len(row)):
                # Convert the value to a number and remove leading zeros
                if is_decimal_string(row[i]):
                    try:
                        row[i] = float(row[i])
                    except ValueError:
                        pass  # Handle the case where the cell doesn't contain a number
            ws.append(row)

    filename_without_extension = os.path.splitext(fileName)[0]

    newFileName = filename_without_extension + excel_save_file_extension
    wb.save(newFileName)

    return newFileName

    
    #Note for future if I need more control over the data going into each row:
    """ 
    with open('classics.csv') as f:
        reader = csv.reader(f, delimiter=',')

        for row_index, row in enumerate(reader, start=1):
            for column_index, cell_value in enumerate(row, start=1):
            ws.cell(row=row_index, column=column_index).value=cell_value
    """

def show_option_dialog(title, question):
    result = messagebox.askquestion(title, question)
    if result == 'yes':
        return True
    else:
        return False

def file_save_as():
    files = [('Excel File', '*.xlsx')]   
            # Not gonna allow saving as CSV since we want it in the fomat of excel anyway.
            #,('CSV File', '*.csv')
    file = asksaveasfile(filetypes = files, defaultextension = files) 
    #with open(file, "wb") as f:
    if file:
        write_to_excel(treev, file)
        return True
    else:
        return False
    

#function to hide and reveal columns
def add_back_column(tree, index):
    global column_width_setting
    global removed_columns_str # this was a string but now is just a heading
    # Remove the second column
    tree.column(treev['columns'][index], width=column_width_setting, stretch=False)
    tree.heading(treev['columns'][index], text=treev['columns'][index])
    #update_sign_in_dropdowns(treev['columns'][index])
    update_columns_in_dropdowns(treev['columns'][index])
    update_date_dropdown(treev['columns'][index])
    update_desc_dropdown(treev['columns'][index])
    update_value_dropdown(treev['columns'][index])

#function to hide and reveal columns along with remove them from the dropdown menus:
def remove_column(tree, index):
    global removed_columns_str
    # Remove the second column
    tree.column(treev['columns'][index], width=0, stretch=False)
    tree.heading(treev['columns'][index], text=removed_columns_str)
    #update_sign_in_dropdowns(treev['columns'][index])
    update_columns_in_dropdowns(treev['columns'][index])
    update_date_dropdown(treev['columns'][index])
    update_desc_dropdown(treev['columns'][index])
    update_value_dropdown(treev['columns'][index])
    
# Callback function to handle the checkbox state change
def checkbox_state_changed(index):
    #global col_storage_pd
    global treev
    """Callback function that gets triggered when a checkbox state changes."""
    state = check_vars[index].get()  # Get the state of the checkbox
    #print(f"Checkbox {index + 1} is {'Checked' if state else 'Unchecked'}")
    index = index + 1 # this is becuase we are ignoring the custom column
    if state:
        #this means that the column should be removed
        #pd.concat([col_storage_pd, copy_column(index)], ignore_index=True)
        remove_column(treev, index)
    else:
        #this means the column should be included
        #move back to treeview
        add_back_column(treev, index)

def cate_add_warnings():
    global cate_add_warnings_bool
    if cate_add_warnings_bool:
        cate_add_warnings_bool = False
    else:
        cate_add_warnings_bool = True

def allow_multiple_file_uploads():
    global allow_multiple_file_uploads_bool
    if allow_multiple_file_uploads_bool:
        allow_multiple_file_uploads_bool = False
    else:
        if messagebox.showwarning("DANGEROUS OPTION!", "Are you sure you want to allow for opening muliple file? This can cause issues with columns and currupt data. Only do this if you are sure that the data which is being opened has all the same columns.", type=messagebox.OKCANCEL):
            allow_multiple_file_uploads_bool = True



def load_file():
    global column_width_setting
    global init_column_width_setting
    global hasLoadedFileData
    global check_vars
    global ButtonsN
    global wb
    file_path = open_file()
    if file_path:
        if not allow_multiple_file_uploads_bool:
            if hasLoadedFileData:
                if show_option_dialog("Save current Data!", "Would you like to first save the data you are working with before you overwrite it with new data?"):
                    if file_save_as():
                        firstTimeOver = False
                    else:
                        return
                # Remove all data in the Treeview
                for item in treev.get_children():
                    treev.delete(item)
        
        #remove all checkboxes:
        for i in range(len(ButtonsN)): # we dont want to use i since it is not the front item thwn we want to 
            #delete first item as we go through the list.
            ButtonsN[0].destroy()
            ButtonsN.pop(0)
                
        hasLoadedFileData = True #this prevents us from loading two different data sets with different columns.
        
        #check if files are CSV then convert to excel:
        file_extension = os.path.splitext(file_path)[1]
        if file_extension == '.CSV':
            file_path = convert_csv_to_excel(file_path)
        
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)

        # Select the active sheet
        sheet = wb.active
        
        firstTimeOver = True
        skip_add_col = -1
        #columnsInTree = ()

        unique_custom_values = []



        # Read and print the data
        for row in sheet.iter_rows(min_row=1, values_only=True):
            row_without_custom = () # this is gonna be used for the options in the dropdown boxes.
            # After the first run through row 1 we will no longer be adding columns
            if firstTimeOver is True:
                # Defining number of columns
                # Find the index of 'Custom Category'
                try:
                    skip_add_col = row.index('Custom Category')
                    temprow = list(row)
                    temprow.pop(0)
                    row_without_custom = temprow
                except ValueError: #if custom was not found then we add the custom column
                    row_without_custom = row
                    row = ("Custom Category",) + row
                
                treev['columns'] = (row)
                column_width_setting = int(int((init_column_width_setting * 8)) / (int(len(treev['columns'])) + 1))
                check_vars = []
                for i,header in enumerate(row):
                    if i == 0: # skip adding the removeal checkbox for custom column
                        continue
                    var = IntVar()
                    
                    check_vars.append(var)
                    tempButt = Checkbutton(label_frame, 
                        text = header, 
                        variable = var, 
                        onvalue = 1, 
                        offvalue = 0, 
                        height = 2, 
                        width = 20,
                        command=lambda index=i-1: checkbox_state_changed(index)) # -1 because we are ignoring the custom column
                    ButtonsN.append(tempButt)

                #Add option to the combo boxes for the selection of categories from the list of headers
                dropdown_date['values'] = row_without_custom
                dropdown_desc['values'] = row_without_custom
                dropdown_value['values'] = row_without_custom
                change_sign_combobox['values'] = row_without_custom
                change_column_combobox['values'] = row_without_custom

                for butt in ButtonsN:
                    butt.pack(side='top', anchor='center')
            else:
                custom_row_without_first_index = list(row)
                if skip_add_col == -1: # if there is not data from an imported file custom column.
                    row = ("",) + tuple(custom_row_without_first_index)
                else:
                    if row[0] == None:
                        temp_list = list(row)
                        temp_list[0] = "" # update none values to blank strings.
                        row = temp_list
                    else:
                        if row[0] not in unique_custom_values:
                            # add to unique values and add it to the combo box for 
                            unique_custom_values.append(row[0])
                            dropdown["values"] = unique_custom_values

                treev.insert("", 'end', text ="L1", values =(row))
            for i,col in enumerate(row):
                if firstTimeOver is True:
                    # Assigning the heading names to the respective columns
                    treev.column(str(i), width=column_width_setting, stretch=False)
                    treev.heading(column=str(i), text=col)
            
            firstTimeOver = False # After the first run through row 1 we will no longer be adding columns

        # Defining heading
        treev['show'] = 'headings'
        dropdown.set('')


#Tree Saving to file:
def write_to_excel(tree, filename):
    """Writes data from the Treeview to an Excel file."""

    # Get the data from the Treeview
    data = []
    for item in tree.get_children():
        values = tree.item(item, "values")
        data.append(values)

    # Create a Pandas DataFrame from the data
    df = pd.DataFrame(data, columns=[tree.heading(col, "text") for col in tree["columns"]])

    if len(removed_columns_str):
        try:
            df.drop(columns=removed_columns_str, axis=1, inplace=True)
        except KeyError:
            print("NO COLUMNS WERE REMOVED")

    # Save the DataFrame to an Excel file
    df.to_excel(filename.name, index=False)



def attempt_end_program():
    if show_option_dialog("Save and Quit?", "Would you like to save the file before you Exit?"):
        if file_save_as():
            messagebox.showinfo("File Saved","File Saved Sucessfully!")
        else:
            messagebox.showerror("FAILED SAVE","File Did Not Save Sucessfully!")
    if messagebox.askokcancel("Quit?", "Would you like to Quit the Program?"):
        r.destroy()
        

r.protocol("WM_DELETE_WINDOW", attempt_end_program)

def about():
    string_ver = "Current Version: " + str(VERSION_MAJOR) + "." + str(VERSION_MINOR) + "." + str(VERSION_PATCH) + "v"
    messagebox.showinfo("About", string_ver)

def Proper_program_usage():
    string_ex = "The Program Intent\n\n" + "The intentbeing the way that this program functions is to allow the user to load in creditcard data and bank data and easily turn them into a graphical representation that can be viewed and explored.\n\n" + "First thing you should do when using this program is upload your bank and creditcard csv or excel files.\n" + "Then remove whatever columns are useless.\n" + "Then add a custom category to the data rows which you want to classify them all under when exported.\n" + "Finally, export the data and select the type of output you want to see when the export is complete.\n\n" + "Please note that when it comes to the Month Averaging or month values it is setup to average over 12 months without setting modification. However, if you turn on the option for month averaging to influce all past months only, it will not include months not vissible in the data. The option for not showing zero values does not affect the average value. This is just to hide data that overfills the screen." 
    messagebox.showinfo("How To...", string_ex)

def changeMonthAveraging():
    pass

#Adding a menu Bar:

def donothing():
    return

menubar = Menu(r)
filemenu = Menu(menubar, tearoff=0)
filemenu.add_command(label="New", command=donothing)
filemenu.add_command(label="Open", command=load_file)
filemenu.add_command(label="Save", command=file_save_as)
filemenu.add_command(label="Export", command=open_export_window)
filemenu.add_separator()
filemenu.add_command(label="Exit", command=attempt_end_program)
menubar.add_cascade(label="File", menu=filemenu)

overrideMenu = Menu(menubar, tearoff=0)
overrideMenu.add_checkbutton(label="Allow Multiple File", command=allow_multiple_file_uploads)
overrideMenu.add_checkbutton(label="Turn Off Category Addition Warnings", command=cate_add_warnings)
menubar.add_cascade(label="OverRide", menu=overrideMenu)

exportMenu = Menu(menubar, tearoff=0)
remove_zero_months_var = tk.BooleanVar()
include_zeros_month_avg_var = tk.BooleanVar()
exportMenu.add_checkbutton(label="Month Averaging: include all previous and current months only.",variable=include_zeros_month_avg_var)
exportMenu.add_checkbutton(label="Month Values: do not include months with zero value in export List.", variable=remove_zero_months_var)
menubar.add_cascade(label="Additional Export Options", menu=exportMenu)


helpmenu = Menu(menubar, tearoff=0)
helpmenu.add_command(label="Help Index", command=Proper_program_usage)
helpmenu.add_command(label="About...", command=about)
menubar.add_cascade(label="Help", menu=helpmenu)

r.config(menu=menubar)


'''
Widgets are added here
'''

r.mainloop()